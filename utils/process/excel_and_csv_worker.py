import os
import openpyxl
import csv
import shutil
import pandas as pd
import calendar

from datetime import datetime, timedelta

from tempfile import NamedTemporaryFile

from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.styles import numbers, Font, Border, Side, PatternFill


fake_db={
    'Батыгин Антон':{
        'target_gp':100163,
        'target_rides':607
    },

    'Бутова Светлана':{
        'target_gp':629764,
        'target_rides':3817
    },

    'Осипова Елена':{
        'target_gp':123915,
        'target_rides':751
    },

    'Кузнецов Дмитрий':{
        'target_gp':78029,
        'target_rides':473
    },

    'Кузнецова Оксана':{
        'target_gp':59400,
        'target_rides':360
    },

    'Шевцов Максим':{
        'target_gp':196070,
        'target_rides':1188
    },

}

koefs = {
        0:1.02, # monday 
        1:1.03, # tuesday
        2:1.13, # wednesday
        3:1.09, # thusrsday 
        4:0.96, # friday
        5:0.94, # saturday
        6:0.81  # sunday
        }


def get_info_from_csv_by_manager_id(id:int, csv_path:str) -> dict:
    id = str(id)
    with open(csv_path, encoding='utf-8') as r_file:
        file_reader = csv.DictReader(r_file, delimiter=',')
        for row in file_reader:
            if row['manager_id'] == id:
                return row


def get_all_info_from_csv(csv_path:str) -> list:
    all_data = list()
    with open(csv_path, encoding='utf-8') as r_file:
        file_reader = csv.DictReader(r_file, delimiter=',')
        for row in file_reader:
            all_data.append(row)

    return(all_data) 


def get_target_rides_info_by_manager_name_from_csv(csv_path:str, manager_name:str):
        df = pd.read_csv(csv_path)
        ds = df[df['full_name']==manager_name]
        return int(ds['target_rides'][0])
        

def get_ratio_info_by_manager_name_from_csv(csv_path:str, manager_name:str):
        df = pd.read_csv(csv_path)
        rides_as_unit_of_ratio = df[df['full_name']==manager_name][rides_as_unit_of_ratio][0]
        return rides_as_unit_of_ratio


def get_manager_id_by_name_from_csv(csv_path:str, manager_name:str) -> int:
    df = pd.read_csv(csv_path)
    ds = df[df['full_name'] == manager_name]
    return int(ds['manager_id'][0])


def get_target_gp_by_name_from_csv(csv_path:str, manager_name:str) -> int:
    df = pd.read_csv(csv_path)
    ds = df[df['full_name'] == manager_name]
    return int(ds['target_gp'][0])


def set_info_to_csv_by_manager_name(csv_path:str, manager_name:str, rides_as_unit_of_ratio=0, targer_rides=0):
    id = get_manager_id_by_name_from_csv(csv_path, manager_name)
    
    old_data = pd.read_csv(csv_path)
    
    targer_rides_to_write = old_data.iloc[id]['target_rides']
    rides_as_unit_of_ratio_to_write = old_data.iloc[id]['rides_as_unit_of_ratio']

    if targer_rides != 0:
        targer_rides_to_write=targer_rides
    
    if rides_as_unit_of_ratio != 0:
        rides_as_unit_of_ratio_to_write = rides_as_unit_of_ratio


    df = pd.read_csv('data_example.csv')
    df.loc[id, 'target_rides'] = targer_rides_to_write
    df.loc[id, 'rides_as_unit_of_ratio'] = rides_as_unit_of_ratio_to_write
    df.to_csv('data_example.csv', index=False)

    print('data for ' + manager_name + ' successfully changed')


def set_info_to_csv_by_manager_id(csv_path:str, id:int, rides_as_unit_of_ratio=0, targer_rides=0):
    old_data = pd.read_csv(csv_path)
    
    targer_rides_to_write = old_data.iloc[id]['target_rides']
    rides_as_unit_of_ratio_to_write = old_data.iloc[id]['rides_as_unit_of_ratio']

    if targer_rides != 0:
        targer_rides_to_write=targer_rides
    
    if rides_as_unit_of_ratio != 0:
        rides_as_unit_of_ratio_to_write = rides_as_unit_of_ratio


    df = pd.read_csv('data_example.csv')
    df.loc[id, 'target_rides'] = targer_rides_to_write
    df.loc[id, 'rides_as_unit_of_ratio'] = rides_as_unit_of_ratio_to_write
    df.to_csv('data_example.csv', index=False)


def create_daily_report_from_file(
    filepath:str='reports/daily_report.xlsx', 
    data_filepath:str='downloads/base/CustomersReports.xlsx', 
    data_for_projection:str='downloads/projection/CustomersReports.xlsx'
        ):

    wb_daily = Workbook()
    wb_daily.save(filepath)
    wb_daily.guess_type = True

    ws_daily = wb_daily.active
    today = datetime.now()
    yesterday = today - timedelta(1)

    ws_daily.append([
        'manager', 
        f'rides {yesterday.day}.{yesterday.month}.{yesterday.year}', 
        'fact rides', 
        'projection', 
        'new rides', 
        'target rides', 
        'fact vs target', 
        'projection vs target',
        'GMV (fact)', 
        'FTR users', 
        'STR users', 
        'total(STR+FTR)', 
        'fact GP', 
        'GP per ride', 
        '%GP', 
        'projection GP', 
        'target GP', 
        'fact vs target (GP)',
        'projection vs Target(GP)'
    ])

    ws_daily.title = 'daily'
    wb_daily.save(filepath)



    
    df = pd.read_excel(data_filepath, 'sheet1')
    proj_df = pd.read_excel(data_for_projection, 'sheet1')

    for manager, target in fake_db.items():
        extra_ds = proj_df[proj_df['Менеджер'] == manager]
        all_ds = proj_df[proj_df['Менеджер'] == manager].sum(numeric_only=True)
        yesterday_ds = df[df['Менеджер'] == manager].sum(numeric_only=True)

        print(manager,'\n',all_ds[[2,4,5,7,8,9]])
        print(manager,'\n',yesterday_ds[[2,4,5,7,8,9]])

        yesterday_rides = yesterday_ds[2]
        all_fact_rides = all_ds[2] 
        
        GMV = yesterday_ds[4]
        target_rides = target.get('target_rides')
        fact_vs_target_rides = (all_fact_rides/target_rides)

        print(manager, all_fact_rides, target_rides, fact_vs_target_rides)

        customers_with_access_to_personal_account = yesterday_ds[7]
        FTR_users = yesterday_ds[8]
        STR_users = yesterday_ds[9]
        total_active_users = FTR_users + STR_users

        margin_GP = all_ds[5]
        GP_per_ride = round(margin_GP/all_fact_rides,2)
        GP_client_check = margin_GP/GMV

        target_gp = target.get('target_gp')


        new_companies_rides = 0
        first_ride_dates = extra_ds.iloc[:, [5,7]]

        for index, row in first_ride_dates.iterrows():
            date_of_first_ride = str(row[0])
            rides_value = row[1]

            try:
                date_of_first_ride = datetime.strptime(date_of_first_ride, '%d.%m.%Y')
            except ValueError:
                continue


            if date_of_first_ride.month == today.month:
                new_companies_rides+=rides_value
                print(manager, date_of_first_ride.month, today.month, )

        # projection
        days_counter = [0] * 7
        last_day_of_month = calendar.monthrange(datetime.now().year, datetime.now().month)[1]+1
        ratio = round(all_fact_rides/(today.day-1), 2)
        summary_coef = 0

        for i in range(today.day, last_day_of_month):
            weekday = calendar.weekday(datetime.now().year, datetime.now().month, i)
            days_counter[weekday] += 1
            summary_coef += koefs[weekday]

        projection_rides = int((ratio*summary_coef)+all_fact_rides)
        fact_vs_projection_rides=all_fact_rides/projection_rides
        projection_gp = projection_rides * GP_per_ride
        projection_vs_target_gp = projection_gp/target_gp
        fact_vs_target_gp = round(margin_GP/target_gp,2)
        projection_vs_target_rides = round(projection_rides/target_rides,2)
        print(manager, all_fact_rides, projection_rides, fact_vs_projection_rides)
        row = [
            manager,
            yesterday_rides,
            all_fact_rides,
            projection_rides,
            new_companies_rides,
            target_rides,
            fact_vs_target_rides,
            projection_vs_target_rides,            
            GMV,
            FTR_users,
            STR_users,
            total_active_users,
            margin_GP,
            GP_per_ride,
            GP_client_check,
            projection_gp,
            target_gp,
            fact_vs_target_gp,
            projection_vs_target_gp
        ]

        ws_daily.append(row)

        wb_daily.save(filepath)


    # total
    ws_daily.cell(row=ws_daily. max_row+2, column=1).value = 'total'

    cols_to_sum = [2,3,4,5,6,7,8,9,10,11,12,13,14,15,16,17]
    row = ws_daily.max_row
    for j in cols_to_sum:
        total = 0
        for i in range(2, ws_daily.max_row):
            try:
                total += int(ws_daily.cell(row=i, column=j).value)
            except:
                continue
        ws_daily.cell(row=row, column=j).value = total


    ws_daily.cell(row=row, column=7).value = float(int(ws_daily.cell(row=row, column=3).value)/int(ws_daily.cell(row=row, column=6).value))
    ws_daily.cell(row=row, column=8).value = float(int(ws_daily.cell(row=row, column=4).value)/int(ws_daily.cell(row=row, column=6).value))
    ws_daily.cell(row=row, column=14).value = int(int(ws_daily.cell(row=row, column=13).value)/int(ws_daily.cell(row=row, column=3).value))
    ws_daily.cell(row=row, column=18).value = float(int(ws_daily.cell(row=row, column=13).value)/int(ws_daily.cell(row=row, column=17).value))
    ws_daily.cell(row=row, column=19).value = float(int(ws_daily.cell(row=row, column=16).value)/int(ws_daily.cell(row=row, column=17).value))



    #set styles

    # columns
    # 
 

    for cell in ws_daily['G']:
        cell.number_format = '0.00%'
    
    for cell in ws_daily['R']:
        cell.number_format = '0.00%'

    for cell in ws_daily['S']:
        cell.number_format = '0.00%'

    for cell in ws_daily['H']:
        cell.number_format = '0.00%'


    for cell in ws_daily['C']:
        cell.number_format = '### ### ##0'

    for cell in ws_daily['F']:
        cell.number_format = '### ### ##0'

    for cell in ws_daily['Q']:
        cell.number_format = '### ### ##0'

    for cell in ws_daily['P']:
        cell.number_format = '### ### ##0'

    for cell in ws_daily['M']:
        cell.number_format = '### ### ##0'

    for cell in ws_daily['I']:
        cell.number_format = '### ### ##0'

    # for cell in ws_daily['M']:
    #     cell.number_format = '0'





 

    # font, styles and width
    ws_daily.column_dimensions['A'].width = '20'
    ws_daily.column_dimensions['B'].width = '18'
    ws_daily.column_dimensions['D'].width = '12'
    ws_daily.column_dimensions['E'].width = '12'
    ws_daily.column_dimensions['F'].width = '15'
    ws_daily.column_dimensions['G'].width = '15'
    ws_daily.column_dimensions['H'].width = '15'
    ws_daily.column_dimensions['I'].width = '15'
    ws_daily.column_dimensions['J'].width = '17'
    ws_daily.column_dimensions['K'].width = '17'
    ws_daily.column_dimensions['L'].width = '15'
    ws_daily.column_dimensions['M'].width = '12'
    ws_daily.column_dimensions['N'].width = '15'
    ws_daily.column_dimensions['O'].width = '20'
    ws_daily.column_dimensions['P'].width = '15'
    ws_daily.column_dimensions['Q'].width = '25'
    ws_daily.column_dimensions['R'].width = '25'
    ws_daily.column_dimensions['S'].width = '25'


    thins = Side(border_style='medium', color='000000')
    # row
    for cell in ws_daily['1']:
        cell.font = Font(bold=True)
    for cell in ws_daily[ws_daily.max_row]:
        cell.font = Font(bold=True)
    for cell in ws_daily[ws_daily.max_row-2]:
        cell.border = Border(bottom=thins)

    wb_daily.save(filepath)


    # Colors
    red_fill = PatternFill(fill_type='solid', fgColor='ff0000')
    yellow_fill = PatternFill(fill_type='solid', fgColor='ffff00')
    green_fill = PatternFill(fill_type='solid', fgColor='008000')

    columns_to_color = ['G', 'H', 'R', 'S']
    for col in columns_to_color:
        for cell in ws_daily[col]:
            try:
                if float(cell.value) < 0.85:
                    cell.fill = red_fill
                if 1 > float(cell.value) >= 0.85:
                    cell.fill = yellow_fill
                if float(cell.value) > 1:
                    cell.fill = green_fill
            except:
                pass
        

    wb_daily.save(filepath)

















        


