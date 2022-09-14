import os
import openpyxl
from openpyxl import load_workbook
from openpyxl.workbook import Workbook

def create_daily_report_from_file(filepath:str='reports/daily_report.xlsx', data_filepath:str='downloads/CustomersReports.xlsx'):
    # if not os.path.isfile(filepath):
    #     wb = Workbook()
    # else:
    #     wb = load_workbook(filepath)

    wb = Workbook()
    ws = wb.active
    ws.title = 'daily'
    wb.save(filepath)

    wb_data = load_workbook(data_filepath)
    ws_data = wb_data.active

    # #получаем всех уникальных менеджеров
    managers = set()
    for i in range(2, ws_data.max_row):
        managers.add(str(ws_data.cell(i, 3).value).strip())
    managers.discard('None')

    list(managers)
    managers = sorted(managers)


    # считаем данные для каждого менеджера
    for line, manager in enumerate(managers,start=1):
        all_rides_for_today = 0
        GMV = 0
        clients_with_access_to_LK = 0
        FTR_users = 0
        STR_users = 0
        GP = 0




        for i in range(2,ws_data.max_row):
            if ws_data.cell(i, 3).value == manager:
                all_rides_for_today+=int(ws_data.cell(i, 8).value)
                if ws_data.cell(i, 11).value is not None:
                    GMV += int(ws_data.cell(i, 11).value)
            else:
                continue

        print(line,manager, all_rides_for_today, GMV)
        ws.cell(line, 2).value = manager
        ws.cell(line, 3).value = all_rides_for_today
        ws.cell(line, 4).value = GMV
    wb.save(filepath)

        
        
